__all__ = [
    'save_cfg',
    'load_cfg',
    'load_frame',
    'load_spatial_mask',
    'save_frame',
    'log_parameters_calculation',
    'log_to_HDU',
    '_LOG_EXPORT_RECARR',
    '_LOG_EXPORT',
    '_LOG_COLUMNS']

import os
import configparser
import logging
import numpy as np
import pandas as pd

from sys import exit
from pathlib import Path
from collections.abc import Sequence

from astropy.io import fits
from .tables import table_fluxes
from . import Error

try:
    import openpyxl
    openpyxl_check = True
    from openpyxl.utils.dataframe import dataframe_to_rows

except ImportError:
    openpyxl_check = False

try:
    import asdf
    asdf_check = True
except ImportError:
    asdf_check = False


try:
    import tomllib
except ModuleNotFoundError:
    import tomli as tomllib

try:
    import toml
    toml_check = True
except ImportError:
    toml_check = False

_logger = logging.getLogger('LiMe')

# Reading file with the format and export status for the measurements
_LIME_FOLDER = Path(__file__).parent
_params_table_file = _LIME_FOLDER/'resources/types_params.txt'
_PARAMS_CONF_TABLE = pd.read_csv(_params_table_file, sep='\s+', header=0, index_col=0)

_LINES_DATABASE_FILE = _LIME_FOLDER/'resources/parent_bands.txt'
_CONF_FILE = _LIME_FOLDER/'config.toml'

# # Read lime configuration file
# with open(_CONF_FILE, mode="rb") as fp:
#     _cfg_lime = tomllib.load(fp)

# Dictionary with the parameter formart
_LOG_COLUMNS = dict(zip(_PARAMS_CONF_TABLE.index.values,
                        _PARAMS_CONF_TABLE.loc[:, 'Norm_by_flux':'dtype'].values))

# Parameters notation latex formatDictionary with the parameter formart
_LOG_COLUMNS_LATEX = dict(zip(_PARAMS_CONF_TABLE.index.values,
                          _PARAMS_CONF_TABLE.loc[:, 'latex_label'].values))

# Array with the parameters to be included in the output log
_LOG_EXPORT = _PARAMS_CONF_TABLE.loc[_PARAMS_CONF_TABLE.Export_log.to_numpy().astype(bool)].index.to_numpy()
_LOG_EXPORT_TYPES = _PARAMS_CONF_TABLE.loc[_PARAMS_CONF_TABLE.Export_log.to_numpy().astype(bool)].dtype.to_numpy()
_LOG_EXPORT_DICT = dict(zip(_LOG_EXPORT, _LOG_EXPORT_TYPES))
_LOG_EXPORT_RECARR = np.dtype(list(_LOG_EXPORT_DICT.items()))

# Attributes from the profile fittings
_ATTRIBUTES_PROFILE = _PARAMS_CONF_TABLE.loc[_PARAMS_CONF_TABLE.Profile_attributes.to_numpy().astype(bool)].index.to_numpy()
_RANGE_PROFILE_FIT = np.arange(_ATTRIBUTES_PROFILE.size)

# Attributes with measurements for log
_ATTRIBUTES_FIT = _PARAMS_CONF_TABLE.loc[_PARAMS_CONF_TABLE.Fit_attributes.to_numpy().astype(bool)].index.to_numpy()
_RANGE_ATTRIBUTES_FIT = np.arange(_ATTRIBUTES_FIT.size)

# Dictionary with the parameter dtypes
_LOG_TYPES_DICT = dict(zip(_PARAMS_CONF_TABLE.index.to_numpy(),
                           _PARAMS_CONF_TABLE.dtype.to_numpy()))


class LiMe_Error(Exception):
    """LiMe exception function"""


def strtobool(val):
    """Convert a string representation of truth to true (1) or false (0).

    True values are 'y', 'yes', 't', 'true', 'on', and '1'; false values
    are 'n', 'no', 'f', 'false', 'off', and '0'.  Raises ValueError if
    'val' is anything else.
    """
    val = val.lower()
    if val in ('y', 'yes', 't', 'true', 'on', '1'):
        return 1
    elif val in ('n', 'no', 'f', 'false', 'off', '0'):
        return 0
    else:
        raise ValueError("invalid truth value %r" % (val,))


def hdu_to_log_df(file_path, page_name):

    with fits.open(file_path) as hdul:
        hdu_log = hdul[page_name].data

    df_log = pd.DataFrame.from_records(data=hdu_log, index='index')
    #
    # # Change 'nan' to np.nan
    # if 'group_label' in df_log:
    #     idcs_nan_str = df_log['group_label'] == 'nan'
    #     df_log.loc[idcs_nan_str, 'group_label'] = np.nan

    # log_df = Table.read(file_path, page_name, character_as_bytes=False).to_pandas()
    # log_df.set_index('index', inplace=True)
    #
    # # Change 'nan' to np.nan
    # if 'group_label' in log_df:
    #     idcs_nan_str = log_df['group_label'] == 'nan'
    #     log_df.loc[idcs_nan_str, 'group_label'] = np.nan

    return df_log


def parse_lime_cfg(toml_cfg, fit_cfg_suffix='_line_fitting'):

    # Convert the configuration entries from the string format if possible
    if fit_cfg_suffix is not None:
        for section, items in toml_cfg.items():
            if section.endswith(fit_cfg_suffix):
                for i_key, i_value in items.items():
                    try:
                        toml_cfg[section][i_key] = format_option_value(i_value, i_key, section)
                    except:
                        _logger.critical(f'Failure to convert entry: "{i_key} = {i_value}" at section [{section}] ')

    return toml_cfg


# Function to load SpecSyzer configuration file
def load_cfg(file_address, fit_cfg_suffix='_line_fitting'):

    """

    This function reads a configuration file with the `toml format <https://toml.io/en/>`_. The text file extension
    must adhere to this format specifications to be successfully read.

    If one of the file sections has the suffix specified by the ``fit_cfg_suffix`` this function will query its items and
    convert the entries to the format expected by `LiMe functions <https://lime-stable.readthedocs.io/en/latest/inputs/n_inputs4_fit_configuration.html>`_.
    The default suffix is "_line_fitting".

    The function will show a critical warning if it fails to convert an item in a ``fit_cfg_suffix`` section.

    :param file_address: Input configuration file address.
    :type file_address: str, pathlib.Path

    :param fit_cfg_suffix: Suffix for LiMe configuration sections. The default value is "_line_fitting".
    :type fit_cfg_suffix:  str

    :return: Parsed configuration data
    :type: dict

    """

    file_path = Path(file_address)

    # Open the file
    if file_path.is_file():

        # Toml file
        with open(file_path, mode="rb") as fp:
            cfg_lime = tomllib.load(fp)

    else:
        raise LiMe_Error(f'The configuration file was not found at: {file_address}')

    # Convert the configuration entries from the string format if possible
    cfg_lime = parse_lime_cfg(cfg_lime, fit_cfg_suffix)

    return cfg_lime


# Function to save SpecSyzer configuration file
def save_cfg(param_dict, output_file, section_name=None, clear_section=False):

    """
    This function safes the input dictionary into a configuration file. If no section is provided the input dictionary
    overwrites the data

    """
    # TODO for line_fitting/models save dictionaries as line
    output_path = Path(output_file)

    if output_path.suffix == '.toml':
        # TODO review convert numpy arrays and floats64
        if toml_check:
            toml_str = toml.dumps(param_dict)
            with open(output_path, "w") as f:
                f.write(toml_str)
        else:
            raise LiMe_Error(f'toml library is not installed. Toml files cannot be saved')

    # Creating a new file (overwritting old if existing)
    else:

        if section_name == None:

            # Check all entries are dictionaries
            values_list = [*param_dict.values()]
            section_check = all(isinstance(x, dict) for x in values_list)
            assert section_check, f'ERROR: Dictionary for {output_file} cannot be converted to configuration file. Confirm all its values are dictionaries'

            output_cfg = configparser.ConfigParser()
            output_cfg.optionxform = str

            # Loop throught he sections and options to create the files
            for section_name, options_dict in param_dict.items():
                output_cfg.add_section(section_name)
                for option_name, option_value in options_dict.items():
                    option_formatted = formatStringOutput(option_value, option_name, section_name)
                    output_cfg.set(section_name, option_name, option_formatted)

            # Save to a text format
            with open(output_file, 'w') as f:
                output_cfg.write(f)

        # Updating old file
        else:

            # Confirm file exists
            file_check = os.path.isfile(output_file)

            # Load original cfg
            if file_check:
                output_cfg = configparser.ConfigParser()
                output_cfg.optionxform = str
                output_cfg.read(output_file)

            # Create empty cfg
            else:
                output_cfg = configparser.ConfigParser()
                output_cfg.optionxform = str

            # Clear section upon request
            if clear_section:
                if output_cfg.has_section(section_name):
                    output_cfg.remove_section(section_name)

            # Add new section if it is not there
            if not output_cfg.has_section(section_name):
                output_cfg.add_section(section_name)

            # Map key values to the expected format and store them
            for option_name, option_value in param_dict.items():
                option_formatted = formatStringOutput(option_value, option_name, section_name)
                output_cfg.set(section_name, option_name, option_formatted)

            # Save to a text file
            with open(output_file, 'w') as f:
                output_cfg.write(f)

    return


def load_frame(fname, page: str = 'FRAME', levels: list = ['id', 'line']):

    """
    This function reads the input ``file_address`` as a pandas dataframe.

    The expected file types are ".txt", ".pdf", ".fits", ".asdf" and ".xlsx". The dataframes expected format is discussed
    on the `line bands <https://lime-stable.readthedocs.io/en/latest/inputs/n_inputs3_line_bands.html>`_ and `measurements <https://lime-stable.readthedocs.io/en/latest/inputs/n_inputs4_fit_configuration.html>`_ documentation.

    For ".fits" and ".xlsx" files the user can provide a page name ``ext`` for the HDU/sheet. The default name is "_LINELOG".

    To reconstruct a `MultiIndex dataframe <https://pandas.pydata.org/docs/user_guide/advanced.html#advanced-hierarchical>`_
    the user needs to specify the ``sample_levels``.

    :param fname: Lines frame file address.
    :type fname: str, Path

    :param page: Name of the HDU/sheet for ".fits"/".xlsx" files. The default value is "_LINELOG".
    :type page: str, optional

    :param levels: Indexes name list for MultiIndex dataframes. The default value is ['id', 'line'].
    :type levels: list, optional

    :return: lines log table
    :rtype: pandas.DataFrame

    """

    # Check file is at path
    if type(fname).__name__ != 'UploadedFile':
        log_path = Path(fname)
        if not log_path.is_file():
            raise LiMe_Error(f'No lines log found at {fname}\n')

        file_name, file_type = log_path.name, log_path.suffix

    else:
        file_name, file_type = fname, 'UploadedFile'


    try:

        # Fits file:
        if file_type == '.fits':
            log = hdu_to_log_df(log_path, page)

        # Excel table
        elif file_type in ['.xlsx' or '.xls']:
            log = pd.read_excel(log_path, sheet_name=page, header=0, index_col=0)

        # ASDF file
        elif file_type == '.asdf':
            with asdf.open(log_path) as af:
                log_RA = af[page]
                log = pd.DataFrame.from_records(log_RA, columns=log_RA.dtype.names)
                log.set_index('index', inplace=True)

                # # Change 'nan' to np.nan
                # idcs_nan_str = log['group_label'] == 'none'
                # log.loc[idcs_nan_str, 'group_label'] = None

        # Text file
        elif file_type == '.txt':
            log = pd.read_csv(log_path, sep='\s+', header=0, index_col=0, comment='#')

        # Uploaded file from streamlit
        elif file_type == 'UploadedFile':
            log = pd.read_csv(file_name, sep='\s+', header=0, index_col=0, comment='#')

        elif file_type == '.csv':
            log = pd.read_csv(log_path, sep=',', header=0, index_col=0, comment='#')

        else:
            _logger.warning(f'File type {file_type} is not recognized. This can cause issues reading the log.')
            log = pd.read_csv(log_path, sep='\s+', header=0, index_col=0)

    except ValueError as e:
        exit(f'\nERROR: LiMe could not open {file_type} file at {log_path}\n{e}')

    # Restore levels if multi-index
    if log.columns.isin(levels).sum() == len(levels):
        log.set_index(levels, inplace=True)

    return log


def save_frame(fname, dataframe, page='FRAME', parameters='all', header=None, column_dtypes=None,
               safe_version=True, **kwargs):

    """

    This function saves the input ``dataframe`` at the ``fname`` provided by the user.

    The accepted extensions are ".txt", ".pdf", ".fits", ".asdf" and ".xlsx".

    For ".fits" and ".xlsx" files the user can provide a page name for the HDU/sheet with the ``ext`` argument.
    The default name is "FRAME".

    The user can specify the ``parameters`` to be saved in the output file.

    For ".fits" files the user can provide a dictionary to add to the ``fits_header``. The user can provide a ``column_dtypes``
    string or dictionary for the output fits file record array. This overwrites LiMe deafult formatting and it must have the
    same columns as the file names.

    :param fname: Lines frame file address.
    :type fname: str, Path

    :param dataframe: Lines dataframe.
    :type dataframe: pandas.DataFrame

    :param parameters: Output parameters list. The default value is "all"
    :type parameters: list

    :param page: Name of the HDU/sheet for ".fits"/".xlsx" files.
    :type page: str, optional

    :param header: Dictionary for ".fits" and ".asdf" file headers.
    :type header: dict, optional

    :param column_dtypes: Conversion variable for the `records array <https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.to_records.html>`.
                          for the output fits file. If a string or type, the data type to store all columns. If a dictionary, a mapping of column
                          names and indices (zero-indexed) to specific data types.
    :type column_dtypes: str, dict, optional

    :param safe_version: Save LiMe version as footnote or page header on the output log. The default value is True.
    :type safe_version: bool, optional

    """

    # Confirm file path exits
    log_path = Path(fname)
    assert log_path.parent.exists(), LiMe_Error(f'- ERROR: Output lines log folder not found: {log_path.parent}')
    file_name, file_type = log_path.name, log_path.suffix

    if len(dataframe.index) > 0:

        # In case of multi-index dataframe
        if isinstance(dataframe.index, pd.MultiIndex):
            log = dataframe.reset_index()
        else:
            log = dataframe

        # Slice the log if the user provides a list of columns
        if parameters != 'all':
            if isinstance(dataframe.index, pd.MultiIndex):
                parameters_list = np.atleast_1d(list(dataframe.index.names) + list(parameters))
            else:
                parameters_list = np.atleast_1d(parameters)

            lines_log = log[parameters_list]

        else:
            lines_log = log

        # Default txt log with the complete information
        if file_type == '.txt':
            with open(log_path, 'wb') as output_file:
                pd.set_option('multi_sparse', False)
                string_DF = lines_log.to_string()

                # Add meta params
                for key, value in kwargs.items():
                    string_DF += f'\n#{key}:{value}'

                output_file.write(string_DF.encode('UTF-8'))

        # CSV
        elif file_type == '.csv':
            with open(log_path, 'wb') as output_file:
                pd.set_option('multi_sparse', False)
                string_DF = lines_log.to_csv(sep=',', na_rep='NaN')

                # Add meta params
                for key, value in kwargs.items():
                    string_DF += f'\n#{key}:{value}'

                output_file.write(string_DF.encode('UTF-8'))

        # Pdf fluxes table
        elif file_type == '.pdf':
            table_fluxes(lines_log, log_path.parent / log_path.stem, header_format_latex=_LOG_COLUMNS_LATEX,
                         lines_notation=log.latex_label.values, **kwargs)

        # Log in a fits format
        elif file_type == '.fits':
            if isinstance(lines_log, pd.DataFrame):

                # Initiate the header
                header = {} if header is None else header

                # Add the meta parameters
                for key, value in kwargs.items():
                    header[key] = value

                lineLogHDU = log_to_HDU(lines_log, ext_name=page, column_dtypes=column_dtypes, header_dict=header)

                if log_path.is_file(): # TODO this strategy is slow for many inputs
                    try:
                        fits.update(log_path, data=lineLogHDU.data, header=lineLogHDU.header, extname=lineLogHDU.name, verify=True)
                    except KeyError:
                        fits.append(log_path, data=lineLogHDU.data, header=lineLogHDU.header, extname=lineLogHDU.name)
                else:
                    hdul = fits.HDUList([fits.PrimaryHDU(), lineLogHDU])
                    hdul.writeto(log_path, overwrite=True, output_verify='fix')

        # Log in excel format
        elif file_type == '.xlsx' or file_type == '.xls':

            # Check openpyxl is installed else leave
            if openpyxl_check:

                # New excel
                if not log_path.is_file():

                    with pd.ExcelWriter(log_path) as writer:
                        lines_log.to_excel(writer, sheet_name=page)

                        if len(kwargs) > 0:
                            sheet_name = f'LiMe_{kwargs["LiMe"]}'
                            df_empty = pd.DataFrame.from_dict(kwargs, orient='index', columns=['values'])
                            df_empty.to_excel(writer, sheet_name=sheet_name, index=True)

                # Updating existing file
                else:
                    wb = openpyxl.load_workbook(log_path)

                    # Remove if existing and create anew
                    if page in wb.sheetnames:
                        wb.remove(wb[page])
                    sheet = wb.create_sheet(page, index=0)

                    # Add data one row at a time
                    for i, row in enumerate(dataframe_to_rows(lines_log, index=True, header=True)):
                        if len(row) > 1: # TOFIX Remove the frozen list logic
                            sheet.append(row)

                    # Save the data
                    wb.save(log_path)

            else:
                _logger.critical(f'openpyxl is not installed. Lines log {fname} could not be saved')

        # Advance Scientific Storage Format
        elif file_type == '.asdf':
            # TODO review this one and add the metadata
            tree = {page: lines_log.to_records(index=True, column_dtypes=_LOG_TYPES_DICT, index_dtypes='<U50')}

            # Create new file
            if not log_path.is_file():
                af = asdf.AsdfFile(tree)
                af.write_to(log_path)

            # Update file
            else:
                with asdf.open(log_path, mode='rw') as af:
                    af.tree.update(tree)
                    af.update()

        # Extension not recognised
        else:
            raise LiMe_Error(f'output extension "{file_type}" was not recognised for file {log_path}')

    # Empty log
    else:
        _logger.info('The output log has no measurements. An output file will not be saved')

    return


def results_to_log(line, log, norm_flux):

    # Loop through the line components
    for i, comp in enumerate(line.list_comps):

        # Add bands wavelengths
        log.at[comp, 'w1'] = line.mask[0]
        log.at[comp, 'w2'] = line.mask[1]
        log.at[comp, 'w3'] = line.mask[2]
        log.at[comp, 'w4'] = line.mask[3]
        log.at[comp, 'w5'] = line.mask[4]
        log.at[comp, 'w6'] = line.mask[5]

        # Treat every line
        for j in _RANGE_ATTRIBUTES_FIT:

            param = _ATTRIBUTES_FIT[j]
            param_value = line.__getattribute__(param)

            # Get component parameter
            if _LOG_COLUMNS[param][3] and (param_value is not None):
                param_value = param_value[i]

            # De-normalize
            if _LOG_COLUMNS[param][0]:
                if param_value is not None:
                    param_value = param_value * norm_flux

            # Just string for particle
            if j == 7:
                param_value = param_value.label

            # Converting None entries to str (9 = group_label)
            if j == 9:
                if param_value is None:
                    param_value = 'none'

            log.at[comp, param] = param_value

    return


def check_file_dataframe(df_variable, variable_type, ext='FRAME', sample_levels=['id', 'line'], copy_input=True):

    if isinstance(df_variable, variable_type):
        if copy_input:
            output = df_variable.copy()
        else:
            output = df_variable

    elif isinstance(df_variable, (str, Path)):
        input_path = Path(df_variable)
        if input_path.is_file():
            output = load_frame(df_variable, page=ext, levels=sample_levels)
        else:
            _logger.warning(f'Lines bands file not found at {df_variable}')
            output = None

    else:
        output = df_variable

    return output


def check_fit_conf(fit_conf, default_key, group_key, group_list=None, fit_cfg_suffix='_line_fitting'):

    # Check that there is an input configuration
    if fit_conf is not None:

        # Check if we have a file
        if not isinstance(fit_conf, dict):
            input_cfg = load_cfg(fit_conf, fit_cfg_suffix)
        else:
            input_cfg = fit_conf.copy()

        # If requested, check that the group/mask configurations are there
        if group_list is not None:

            for mask_name in group_list:

                mask_fit_cfg = input_cfg.get(f'{mask_name}_line_fitting')

                missing_mask = False
                if mask_fit_cfg is not None:
                    if mask_fit_cfg.get('bands') is None:
                        missing_mask = True
                else:
                    missing_mask = True

                if missing_mask:
                    error_message = 'No input "bands" provided. In this case you need to include the \n' \
                                    f'you need to specify an "bands=log_file_address" entry the ' \
                                    f'"[{mask_name}_file]" of your fitting configuration file'
                    raise LiMe_Error(error_message)

        # Recover the configuration expected for the object
        default_cfg = input_cfg.get(f'{default_key}_line_fitting') if default_key is not None else None
        mask_cfg = input_cfg.get(f'{group_key}_line_fitting') if group_key is not None else None

        # Case there are not leveled entries
        if (default_cfg is None) and (mask_cfg is None):
            output_cfg = input_cfg

        # Proceed to update the levels
        else:

            # Default configuration
            output_cfg = {} if default_cfg is None else default_cfg
            default_detect = output_cfg.get('line_detection')

            # Mask conf
            mask_conf = {} if mask_cfg is None else mask_cfg
            mask_detect = mask_conf.get('line_detection')

            # Update the levels
            output_cfg = {**output_cfg, **mask_conf}

            # If no line detection don't add it
            if mask_detect is not None:
                output_cfg['line_detection'] = mask_detect
            elif default_detect is not None:
                output_cfg['line_detection'] = default_detect
            else:
                pass

    else:
        output_cfg = {}

    return output_cfg


_parent_bands_file = Path(__file__).parent/'resources/parent_bands.txt'
_PARENT_BANDS = load_frame(_parent_bands_file)


def check_numeric_value(s):

    # Function to check if variable can be converte to float else leave as string

    try:
        output = float(s)
        return output
    except ValueError:
        return s


# Function to map a string to its variable-type
def format_option_value(entry_value, key_label, section_label='', float_format=None, nan_format='nan'):

    output_variable = entry_value

    # Dictionary blended lines
    if isinstance(entry_value, str):

        output_variable = {}

        try:
            keys_and_values = entry_value.split(',')
            for pair in keys_and_values:

                # Conversion for parameter class atributes
                if ':' in pair:
                    key, value = pair.split(':')
                    if value == 'None':
                        output_variable[key] = None
                    elif key in ['value', 'min', 'max']:
                        output_variable[key] = float(value)
                    elif key == 'vary':
                        output_variable[key] = strtobool(value) == 1
                    else:
                        output_variable[key] = value
                elif '_mask' in key_label:
                    output_variable = entry_value
                # Conversion for non-parameter class atributes (only str conversion possible)
                else:
                    output_variable = check_numeric_value(entry_value)
        except:
            raise LiMe_Error(f'Failure to convert configuration entry: {key_label} = {entry_value} in section {section_label}')

    return output_variable


def log_parameters_calculation(input_log, parameter_list, formulae_list):

    # Load the log if necessary file:
    file_check = False
    if isinstance(input_log, pd.DataFrame):
        log_df = input_log

    elif isinstance(input_log, (str, Path)):
        file_check = True
        log_df = load_frame(input_log)

    else:
        _logger.critical(
            f'Not a recognize log format. Please use a pandas dataframe or a string/Path object for file {input_log}')
        exit(1)

    # Parse the combined expression
    expr = ''
    for col, formula in zip(parameter_list, formulae_list):
        expr += f'{col}={formula}\n'

    # Compute the new parameters
    log_df.eval(expr=expr, inplace=True)

    # Save to the previous location
    if file_check:
        save_frame(log_df, input_log)


    return



def log_to_HDU(log, ext_name=None, column_dtypes=None, header_dict=None):

    # For non empty logs
    if not log.empty:

        if column_dtypes is None:
            column_dtypes = _LOG_TYPES_DICT

        if header_dict is None:
            header_dict = {}

        linesSA = log.to_records(index=True, column_dtypes=column_dtypes, index_dtypes='<U50')
        linesCol = fits.ColDefs(linesSA)

        hdr = fits.Header(header_dict) if isinstance(header_dict, dict) else header_dict
        linesHDU = fits.BinTableHDU.from_columns(linesCol, name=ext_name, header=hdr)

    # Empty log
    else:
        linesHDU = None

    return linesHDU


def formatStringOutput(value, key, section_label=None, float_format=None, nan_format='nan'):

    # TODO this one should be the default option
    # TODO add more cases for dicts
    # Check None entry
    if value is not None:

        # Check string entry
        if isinstance(value, str):
            formatted_value = value

        else:

            # Case of an array
            scalarVariable = True
            if isinstance(value, (Sequence, np.ndarray)):

                # Confirm is not a single value array
                if len(value) == 1:
                    value = value[0]

                # Case of an array
                else:
                    scalarVariable = False
                    formatted_value = ','.join([str(item) for item in value])

            if scalarVariable:

                # Case single float
                if isinstance(value, str):
                    formatted_value = value
                else:
                    if np.isnan(value):
                        formatted_value = nan_format
                    else:
                        formatted_value = str(value)

    else:
        formatted_value = 'None'

    return formatted_value


# def progress_bar(i, i_max, post_text, n_bar=10):
#
#     # Size of progress bar
#     j = i/i_max
#     stdout.write('\r')
#     message = f"[{'=' * int(n_bar * j):{n_bar}s}] {int(100 * j)}% {post_text}"
#     stdout.write(message)
#     stdout.flush()
#
#     return





def load_spatial_mask(mask_file, mask_list=None, return_coords=False):

    # Masks array container
    spatial_mask_dict = {}

    # Limit the analysis to some masks
    mask_list_check = False if mask_list is None else True

    # Check that mask is there:
    if mask_file is not None:

        mask_file = Path(mask_file)

        if mask_file.is_file():

            with fits.open(mask_file) as maskHDULs:

                # Save the fits data to restore later
                counter = 0
                for HDU in maskHDULs:
                    ext_name = HDU.name
                    if HDU.name != 'PRIMARY':
                        if mask_list_check:
                            if ext_name in mask_list:
                                spatial_mask_dict[ext_name] = (HDU.data.astype('bool'), HDU.header)
                                counter += 1
                        else:
                            spatial_mask_dict[ext_name] = (HDU.data.astype('bool'), HDU.header)
                            counter += 1

                # Warn if the mask file does not contain any of the expected masks
                if counter == 0:
                    _logger.warning(f'No masks extensions were found in file {mask_file}')

        else:
            _logger.warning(f'No mask file was found at {mask_file.as_posix()}.')

    # Return the mask as a set of coordinates (no headers)
    if return_coords:
        for mask_name, mask_data in spatial_mask_dict.items():
            idcs_spaxels = np.argwhere(mask_data[0])
            spatial_mask_dict[mask_name] = idcs_spaxels

    return spatial_mask_dict


def check_file_array_mask(var, mask_list=None):

    # Check if file
    if isinstance(var, (str, Path)):

        input = Path(var)
        if input.is_file():
            mask_dict = load_spatial_mask(var, mask_list)
        else:
            raise Error(f'No spatial mask file at {Path(var).as_posix()}')

    # Array
    elif isinstance(var, (np.ndarray, list)):

        # Re-adjust the variable
        var = np.ndarray(var, ndmin=3)
        masks_array = np.squeeze(np.array_split(var, var.shape[0], axis=0))

        # Confirm boolean array
        if masks_array.dtype != bool:
            _logger.warning(f'The input mask array should have a boolean variables (True/False)')

        # Incase user gives a str
        mask_list = [mask_list] if isinstance(mask_list, str) else mask_list

        # Check if there is a mask list
        if len(mask_list) == 0:
            mask_list = [f'SPMASK{i}' for i in range(masks_array.shape[0])]

        # Case the number of masks names and arrays is different
        elif masks_array.shape[0] != len(mask_list):
            _logger.warning(f'The number of input spatial mask arrays is different than the number of mask names')

        # Everything is fine
        else:
            mask_list = mask_list

        # Create mask dict with empty headers
        mask_dict = dict(zip(mask_list, (masks_array, {})))

    else:

        raise Error(f'Input mask format {type(input)} is not recognized for a mask file. Please declare a fits file, a'
                    f' numpy array or a list/array of numpy arrays')

    return mask_dict




