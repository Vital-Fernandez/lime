import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def save_dataframe_to_excel(output_file, df, page_name):
    # Check if the file exists
    file_exists = os.path.exists(output_file)

    # If the file does not exist, save the DataFrame as a new Excel file
    if not file_exists:
        df.to_excel(output_file, index=False, sheet_name=page_name)
        return

    # If the file exists, load the workbook
    workbook = load_workbook(output_file)

    # If the sheet with the same name exists, replace the data
    if page_name in workbook.sheetnames:
        workbook.remove(workbook[page_name])
        sheet = workbook.create_sheet(page_name)
    else:
        # If the sheet with the same name does not exist, create a new sheet
        sheet = workbook.create_sheet(page_name)

    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)

    # Save the updated workbook
    workbook.save(output_file)

# Example usage:
data = {
    'Name': ['John', 'Alice', 'Bob', 'Emma'],
    'Age': [30, 25, 40, 35],
    'City': ['New York', 'Los Angeles', 'Chicago', 'Houston']
}
df = pd.DataFrame(data)

# Call the function with your desired file address, DataFrame, and page name
save_dataframe_to_excel("output.xlsx", df, "Sheet1")
save_dataframe_to_excel("output.xlsx", df, "Sheet2")
df['Age'] = df['Age'] * 2
save_dataframe_to_excel("output.xlsx", df, "Sheet2")

